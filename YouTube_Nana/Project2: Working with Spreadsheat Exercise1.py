import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file [ "product_list" ]

product_per_supplier = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value

    if supplier_name in product_per_supplier:
        product_per_supplier [ supplier_name ] = product_per_supplier [ supplier_name ] + 1
    else:
        product_per_supplier [ supplier_name ] = 1

print(product_per_supplier)
