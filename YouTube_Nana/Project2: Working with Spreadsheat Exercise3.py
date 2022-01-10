import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["product_list"]

products_under10_inventory = {}

for product_row in range(2, product_list.max_row + 1):

    # supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value

    if inventory < 10:
        products_under10_inventory[int(product_number)] = int(inventory)

print(products_under10_inventory)
