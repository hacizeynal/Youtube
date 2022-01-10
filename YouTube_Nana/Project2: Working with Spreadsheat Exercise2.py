import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["product_list"]

total_value_per_supplier = {}

for product_row in range(2, product_list.max_row + 1):
    #exceldeki column-lari variable olaraq assign ele

    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    if  supplier_name in total_value_per_supplier:
        current_value = total_value_per_supplier[supplier_name]
        total_value_per_supplier[supplier_name] = current_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

print(total_value_per_supplier)
