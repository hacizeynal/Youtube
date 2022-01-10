import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["product_list"]

products_under10_inventory = {}

for product_row in range(2, product_list.max_row + 1):

    # supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inventory_cost = product_list.cell(product_row, 5)

    inventory_cost.value = price * inventory

inventory_file.save("updated_inventory_file.xlsx")
